import zipfile
import os
import pytest
from PyPDF2 import PdfReader
from openpyxl import load_workbook

DIRECTORY = 'resources'
ARCHIVE_PATH = f'{DIRECTORY}/archive.zip'
XLSX_PATH = f'{DIRECTORY}/example.xlsx'
PDF_PATH = f'{DIRECTORY}/file.pdf'
CSV_PATH = f'{DIRECTORY}/username.csv'


@pytest.fixture()
def create_zip():
    try:
        os.remove(ARCHIVE_PATH)
    except FileNotFoundError:
        pass
    finally:
        file_dir = os.listdir(DIRECTORY)
        with zipfile.ZipFile(ARCHIVE_PATH, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
            for file in file_dir:
                add_file = os.path.join(DIRECTORY, file)
                zf.write(add_file)


def define_xlsx_properties(file):
    workbook = load_workbook(file)
    sheet = workbook.active
    row_count = sheet.max_row
    column_count = sheet.max_column
    header = [sheet.cell(row=3, column=column_index).value for column_index in range(1, column_count)]
    return row_count, column_count, header


def define_pdf_properties(file):
    reader = PdfReader(file)
    page_count = len(reader.pages)
    text = reader.pages[0].extract_text()
    header = text.split('\n')
    return page_count, header


def test_csv(create_zip):
    with open(CSV_PATH, 'r') as csv_file:
        csv_row = len(list(csv_file))
    with zipfile.ZipFile(ARCHIVE_PATH, 'r') as zf:
        with zf.open(CSV_PATH) as archive_csv_file:
            archive_csv_row = len(list(archive_csv_file))
    assert archive_csv_row == csv_row


def test_pdf(create_zip):
    pdf_pages, pdf_header = define_pdf_properties(PDF_PATH)
    with zipfile.ZipFile(ARCHIVE_PATH, 'r') as zf:
        with zf.open(PDF_PATH) as archive_pdf_file:
            archive_pdf_pages, archive_pdf_header = define_pdf_properties(archive_pdf_file)
    assert archive_pdf_pages == pdf_pages
    assert archive_pdf_header == pdf_header


def test_xlsx(create_zip):
    xlsx_row, xlsx_column, xlsx_header = define_xlsx_properties(XLSX_PATH)
    with zipfile.ZipFile(ARCHIVE_PATH, 'r') as zf:
        with zf.open(XLSX_PATH) as archive_xlsx_file:
            archive_xlsx_row, archive_xlsx_column, archive_xlsx_header = define_xlsx_properties(archive_xlsx_file)
    assert archive_xlsx_row == xlsx_row
    assert archive_xlsx_column == xlsx_column
    assert archive_xlsx_header == xlsx_header
